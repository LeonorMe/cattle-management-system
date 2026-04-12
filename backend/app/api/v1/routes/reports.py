import io
import csv
from typing import Any, List, Optional
from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
import openpyxl
from datetime import date

from app.api import deps
from app.models.animal import Animal
from app.models.event import Event
from app.models.user import User

router = APIRouter()

@router.get("/herd", status_code=200)
def export_herd(
    db: Session = Depends(deps.get_db),
    current_user: User = Depends(deps.get_current_user),
    format: str = Query("xlsx", description="Format: xlsx or csv")
) -> StreamingResponse:
    """Export all animals in the farm."""
    if not current_user.farm_id:
        raise HTTPException(status_code=400, detail="Not assigned to a farm")

    animals = db.query(Animal).filter(Animal.farm_id == current_user.farm_id).all()
    
    headers = ["ID Registro", "Nome", "Raça", "Género", "Data Nasc.", "Estado", "Mãe", "Pai"]

    if format == "csv":
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(headers)
        for a in animals:
            writer.writerow([
                a.registration_id, a.name or "", a.breed or "", 
                a.gender.value, a.birth_date or "", a.status.value,
                a.mother_id or "", a.father_id or ""
            ])
        output.seek(0)
        return StreamingResponse(
            iter([output.getvalue()]),
            media_type="text/csv",
            headers={"Content-Disposition": "attachment; filename=herd_report.csv"}
        )
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Herd"
        ws.append(headers)
        for a in animals:
            ws.append([
                a.registration_id, a.name or "", a.breed or "", 
                a.gender.value, str(a.birth_date) if a.birth_date else "", a.status.value,
                a.mother_id or "", a.father_id or ""
            ])
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=herd_report.xlsx"}
        )

@router.get("/events", status_code=200)
def export_events(
    start_date: Optional[date] = Query(None),
    end_date: Optional[date] = Query(None),
    db: Session = Depends(deps.get_db),
    current_user: User = Depends(deps.get_current_user),
    format: str = Query("xlsx")
) -> StreamingResponse:
    """
    Export events within a date range to Excel or CSV.
    """
    if not current_user.farm_id:
        raise HTTPException(status_code=400, detail="Not assigned to a farm")

    animal_ids = [a.id for a in db.query(Animal.id).filter(Animal.farm_id == current_user.farm_id).all()]
    
    query = db.query(Event).filter(Event.animal_id.in_(animal_ids))
    if start_date:
        query = query.filter(Event.event_date >= start_date)
    if end_date:
        query = query.filter(Event.event_date <= end_date)
        
    events = query.order_by(Event.event_date.desc()).all()
    headers = ["Data", "Animal (ID)", "Tipo de Evento", "Descrição"]

    if format == "csv":
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(headers)
        for e in events:
            animal = db.query(Animal).filter(Animal.id == e.animal_id).first()
            writer.writerow([
                e.event_date, animal.registration_id if animal else "N/A", 
                e.event_type.value, e.description or ""
            ])
        output.seek(0)
        return StreamingResponse(
            iter([output.getvalue()]),
            media_type="text/csv",
            headers={"Content-Disposition": "attachment; filename=events_report.csv"}
        )
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Events"
        ws.append(headers)
        for e in events:
            animal = db.query(Animal).filter(Animal.id == e.animal_id).first()
            ws.append([
                str(e.event_date), animal.registration_id if animal else "N/A", 
                e.event_type.value, e.description or ""
            ])
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=events_report.xlsx"}
        )

@router.post("/import/animals", status_code=200)
def import_animals(
    file: UploadFile = File(...),
    db: Session = Depends(deps.get_db),
    current_user: User = Depends(deps.get_current_user),
) -> Any:
    """
    Import or update animals from an Excel or CSV file.
    """
    if not current_user.farm_id:
        raise HTTPException(status_code=400, detail="Not assigned to a farm")
        
    from app.models.animal import AnimalGender, AnimalStatus
    from datetime import datetime
    
    content = file.file.read()
    filename = file.filename.lower()
    
    rows = []
    try:
        if filename.endswith(".csv"):
            text = content.decode("utf-8")
            reader = csv.DictReader(io.StringIO(text))
            for row in reader:
                rows.append(row)
        elif filename.endswith(".xlsx"):
            wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
            ws = wb.active
            headers = [cell.value for cell in ws[1]]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or not row[0]: continue  
                row_dict = dict(zip(headers, row))
                rows.append(row_dict)
        else:
            raise HTTPException(status_code=400, detail="Invalid file type. Only .xlsx and .csv are supported.")
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Error parsing file: {str(e)}")
        
    def resolve_parent_id(reg_id):
        if not reg_id: return None
        parent = db.query(Animal).filter(
            Animal.farm_id == current_user.farm_id, 
            Animal.registration_id == str(reg_id).strip()
        ).first()
        return parent.id if parent else None
        
    imported_count = 0
    updated_count = 0
    
    for row in rows:
        reg_id = str(row.get("registration_id", "")).strip()
        if not reg_id: continue
        
        name = row.get("name")
        breed = row.get("breed")
        gender_str = str(row.get("gender", "F")).strip().upper()
        gender = AnimalGender.MALE if gender_str == "M" else AnimalGender.FEMALE
        
        birth_date_val = row.get("birth_date")
        birth_date = None
        if birth_date_val:
            if isinstance(birth_date_val, datetime):
                birth_date = birth_date_val.date()
            else:
                try:
                    birth_date = datetime.strptime(str(birth_date_val).strip(), "%Y-%m-%d").date()
                except ValueError:
                    pass
                    
        status_str = str(row.get("status", "Active")).strip().capitalize()
        status = AnimalStatus.ACTIVE
        if status_str == "Sold": status = AnimalStatus.SOLD
        elif status_str == "Deceased": status = AnimalStatus.DECEASED
        
        mother_reg_id = row.get("mother_registration_id")
        father_reg_id = row.get("father_registration_id")
        
        mother_id = resolve_parent_id(mother_reg_id)
        father_id = resolve_parent_id(father_reg_id)
        
        existing = db.query(Animal).filter(Animal.farm_id == current_user.farm_id, Animal.registration_id == reg_id).first()
        if existing:
            existing.name = name
            existing.breed = breed
            existing.gender = gender
            existing.birth_date = birth_date
            existing.status = status
            existing.mother_id = mother_id
            existing.father_id = father_id
            updated_count += 1
        else:
            new_animal = Animal(
                farm_id=current_user.farm_id,
                registration_id=reg_id,
                name=name,
                breed=breed,
                gender=gender,
                birth_date=birth_date,
                status=status,
                mother_id=mother_id,
                father_id=father_id
            )
            db.add(new_animal)
            imported_count += 1
            
    db.commit()
    return {"status": "success", "imported": imported_count, "updated": updated_count}
